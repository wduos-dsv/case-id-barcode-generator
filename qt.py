import sys
import os
import re
from PyQt6.QtWidgets import (QApplication, QWidget, QVBoxLayout, QPushButton, 
                             QFileDialog, QLabel, QMessageBox, QHBoxLayout, QComboBox)
from PyQt6.QtCore import Qt

# --- Import processing libraries ---
import openpyxl
from openpyxl.drawing.image import Image
import barcode
from barcode.writer import ImageWriter
from PIL import Image as PILImage

class BarcodeApp(QWidget):
    def __init__(self):
        super().__init__()
        self.selected_file_path = ""
        self.initUI()
        
    def initUI(self):
        self.setWindowTitle("Case ID Barcode Generator")
        self.setMinimumSize(480, 220)
        
        layout = QVBoxLayout()
        layout.setSpacing(15)
        
        file_layout = QHBoxLayout()
        self.file_label = QLabel("Nenhum arquivo selecionado...")
        self.file_label.setStyleSheet("color: #666; font-style: italic;")
        
        btn_browse = QPushButton("Procurar...")
        btn_browse.clicked.connect(self.open_file_dialog)
        
        file_layout.addWidget(self.file_label, stretch=4)
        file_layout.addWidget(btn_browse, stretch=1)
        layout.addLayout(file_layout)
        
        sort_layout = QHBoxLayout()
        sort_label = QLabel("Ordenar por:")
        sort_label.setStyleSheet("font-weight: bold;")
        
        self.sort_dropdown = QComboBox()
        self.sort_dropdown.addItems(["SKU", "Endereço", "Tipo de Caixa"])
        
        sort_layout.addWidget(sort_label, stretch=1)
        sort_layout.addWidget(self.sort_dropdown, stretch=4)
        layout.addLayout(sort_layout)
        
        self.btn_process = QPushButton("Iniciar Processamento")
        self.btn_process.setEnabled(False) 
        self.btn_process.setStyleSheet("""
            QPushButton { background-color: #0078d4; color: white; font-weight: bold; padding: 8px; border-radius: 4px; }
            QPushButton:disabled { background-color: #cccccc; color: #666666; }
            QPushButton:hover { background-color: #005a9e; }
        """)
        self.btn_process.clicked.connect(self.process_spreadsheet)
        layout.addWidget(self.btn_process)
        
        self.setLayout(layout)
        
    def open_file_dialog(self):
        file_filter = "Excel Files (*.xlsx *.xlsm)"
        file_path, _ = QFileDialog.getOpenFileName(self, "Selecionar Planilha", "", file_filter)
        
        if file_path:
            self.selected_file_path = file_path
            display_name = os.path.basename(file_path)
            self.file_label.setText(f"Arquivo: {display_name}")
            self.file_label.setStyleSheet("color: #000; font-style: normal; font-weight: bold;")
            self.btn_process.setEnabled(True)

    def process_spreadsheet(self):
        if not self.selected_file_path:
            return
            
        try:
            wb = openpyxl.load_workbook(self.selected_file_path, data_only=True)
            if 'Imprimir' not in wb.sheetnames:
                raise ValueError("A aba 'Imprimir' não foi encontrada no arquivo selecionado.")
                
            ws = wb['Imprimir']
            
            # --- CELL & BARCODE SIZING CONFIGURATION ---
            # 1. Expand Excel Cell dimensions to leave ample room for borders
            ROW_HEIGHT_PT = 45          # Larger row height (45pt)
            COL_WIDTH_CHAR = 22         # Larger column width (22 chars)
            
            CELL_WIDTH_PX = int(COL_WIDTH_CHAR * 7.1)    # ~156px
            CELL_HEIGHT_PX = int(ROW_HEIGHT_PT * 1.333)  # ~60px
            
            # 2. Keep barcode image size fixed to original compact dimensions
            BARCODE_WIDTH_PX = 107
            BARCODE_HEIGHT_PX = 34
            # -------------------------------------------------------------
            
            full_pallets_removed = 0
            rows_data = []
            
            # Filter rows
            for row in range(4, 102):
                row_values = [ws.cell(row=row, column=col).value for col in range(1, 7)]
                if row_values[1] is not None:
                    qty_value = row_values[2]
                    if qty_value not in (360, 400, "360", "400"):
                        rows_data.append(row_values)
                    else:
                        full_pallets_removed += 1
            
            # Sort rows
            sort_method = self.sort_dropdown.currentIndex()
            if sort_method == 0:
                rows_data.sort(key=lambda r: int(r[1]) if str(r[1]).isdigit() else 999999)
            elif sort_method == 1:
                def get_location_key(r):
                    val = str(r[0]) if r[0] is not None else ""
                    parts = re.split(r'(\d+)', val)
                    return [int(text) if text.isdigit() else text.lower() for text in parts]
                rows_data.sort(key=get_location_key)
            elif sort_method == 2:
                rows_data.sort(key=lambda r: str(r[3]).lower() if r[3] is not None else "")
            
            # Clear grid
            for row in range(4, 102):
                for col in range(1, 7):
                    ws.cell(row=row, column=col).value = None
                    
            # Populate sorted data
            for index, row_values in enumerate(rows_data):
                current_row = 4 + index
                for col_idx, val in enumerate(row_values):
                    ws.cell(row=current_row, column=col_idx + 1).value = val
            
            ws.column_dimensions['F'].width = COL_WIDTH_CHAR
            
            # Generate and add centered barcodes
            for index in range(len(rows_data)):
                row = 4 + index
                cell_value = ws[f'E{row}'].value
                
                if cell_value and not str(cell_value).startswith('='):
                    # 1. Generate raw barcode image
                    code128 = barcode.get('code128', str(cell_value), writer=ImageWriter())
                    raw_filename = f"temp_raw_{row}"
                    code128.save(raw_filename, options={'write_text': False, 'module_margin': 1})
                    
                    # 2. Resize barcode to compact size
                    raw_img = PILImage.open(f"{raw_filename}.png")
                    barcode_resized = raw_img.resize((BARCODE_WIDTH_PX, BARCODE_HEIGHT_PX), PILImage.Resampling.LANCZOS)
                    
                    # 3. Create a TRANSPARENT canvas (RGBA with 0 alpha) matching cell size
                    # This prevents solid white pixels from obscuring Excel's cell borders
                    canvas = PILImage.new('RGBA', (CELL_WIDTH_PX, CELL_HEIGHT_PX), (0, 0, 0, 0))
                    
                    paste_x = (CELL_WIDTH_PX - BARCODE_WIDTH_PX) // 2
                    paste_y = (CELL_HEIGHT_PX - BARCODE_HEIGHT_PX) // 2
                    
                    canvas.paste(barcode_resized, (paste_x, paste_y))
                    
                    centered_filename = f"temp_barcode_{row}.png"
                    canvas.save(centered_filename, format="PNG")
                    
                    # 4. Add image to cell safely
                    ws.row_dimensions[row].height = ROW_HEIGHT_PT
                    img = Image(centered_filename)
                    ws.add_image(img, f'F{row}')
            
            # Remove other sheets safely
            for sheet_name in list(wb.sheetnames):
                if sheet_name != 'Imprimir':
                    wb.remove(wb[sheet_name])
            
            output_path = os.path.join(os.path.dirname(self.selected_file_path), 'Imprimir.xlsx')
            wb.save(output_path)
            
            # Cleanup temp files
            for index in range(len(rows_data)):
                row = 4 + index
                for f in [f"temp_raw_{row}.png", f"temp_barcode_{row}.png"]:
                    if os.path.exists(f):
                        os.remove(f)
                    
            QMessageBox.information(self, "Sucesso", f"Processamento concluído com sucesso!\nSalvo em: {output_path}\n\nNúmero de pallets FULL neste pedido: {full_pallets_removed}")
            
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Ocorreu um erro: {str(e)}")

if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = BarcodeApp()
    ex.show()
    sys.exit(app.exec())