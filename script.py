import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU
import barcode
from barcode.writer import ImageWriter
import os

# 1. Load your Excel workbook
wb = openpyxl.load_workbook('Picking by Case ID.xlsx', data_only=True)
ws = wb['Imprimir']

# 2. Setup Barcode options
options = {
    'module_width': 0.2,
    'module_height': 7.0,
    'quiet_zone': 2.0,
    'write_text': False 
}

# Define final layout dimensions once
BARCODE_WIDTH_PX = 130
BARCODE_HEIGHT_PX = 40
ROW_HEIGHT_PT = 45
COL_WIDTH_CHAR = 22

CELL_WIDTH_PX = COL_WIDTH_CHAR * 7.1 
CELL_HEIGHT_PX = ROW_HEIGHT_PT * 1.333

# -------------------------------------------------------------
# STEP A: EXTRACT, FILTER, AND SORT DATA IN MEMORY
# -------------------------------------------------------------
rows_data = []

# Read rows 4 to 101
for row in range(4, 102):
    # Grab values for all columns we care about in this row (A to F)
    row_values = [ws.cell(row=row, column=col).value for col in range(1, 7)]
    
    # Check if the row actually has an Item value in Column B (Index 1)
    if row_values[1] is not None:
        qty_value = row_values[2] # Column C is index 2
        
        # Rule 2: Remove lines where Quantity (Col C) is 400 or 360
        if qty_value not in (360, 400, "360", "400"):
            rows_data.append(row_values)

# Rule 1: Put Column B (Index 1) in numeric order
# We use a lambda to ensure it converts to integer for mathematically accurate sorting
def get_sort_key(r):
    try:
        return int(r[1])
    except (ValueError, TypeError):
        return 999999  # Pushes non-numeric items to the bottom

rows_data.sort(key=get_sort_key)

# -------------------------------------------------------------
# STEP B: CLEAR OLD ROW DATA AND WRITE BACK SORTED DATA
# -------------------------------------------------------------
# Clear data in the original grid area (Rows 4 to 101, Columns A to F)
for row in range(4, 102):
    for col in range(1, 7):
        ws.cell(row=row, column=col).value = None

# Write our sorted/filtered data back into the sheet starting at row 4
for index, row_values in enumerate(rows_data):
    current_row = 4 + index
    for col_idx, val in enumerate(row_values):
        ws.cell(row=current_row, column=col_idx + 1).value = val

# -------------------------------------------------------------
# STEP C: GENERATE BARCODES ON THE NEW CLEAN DATA
# -------------------------------------------------------------
ws.column_dimensions['F'].width = COL_WIDTH_CHAR

# Loop dynamically through only the rows that now contain data
for index in range(len(rows_data)):
    row = 4 + index
    cell_value = ws[f'E{row}'].value
    
    if cell_value and not str(cell_value).startswith('='):
        # Generate the temporary barcode image
        code128 = barcode.get('code128', str(cell_value), writer=ImageWriter())
        filename = f"temp_barcode_{row}"
        code128.save(filename, options=options)
        
        # Insert the image into Column F
        img_path = f"{filename}.png"
        img = Image(img_path)
        
        # Match the Excel row dimension
        ws.row_dimensions[row].height = ROW_HEIGHT_PT
        
        # Centering logic
        left_margin_px = max(0, (CELL_WIDTH_PX - BARCODE_WIDTH_PX) / 2)
        top_margin_px = max(0, (CELL_HEIGHT_PX - BARCODE_HEIGHT_PX) / 2)
        left_margin_emu = int(left_margin_px * 9525)
        top_margin_emu = int(top_margin_px * 9525)
        
        marker = AnchorMarker(col=5, colOff=left_margin_emu, row=row-1, rowOff=top_margin_emu)
        barcode_size = XDRPositiveSize2D(
            cx=pixels_to_EMU(BARCODE_WIDTH_PX), 
            cy=pixels_to_EMU(BARCODE_HEIGHT_PX)
        )
        
        img.anchor = OneCellAnchor(_from=marker, ext=barcode_size)
        ws.add_image(img)

# 5. Remove other sheets
for sheet_name in wb.sheetnames:
    if sheet_name != 'Imprimir':
        sheet_to_delete = wb[sheet_name]
        wb.remove(sheet_to_delete)

# 6. Save the file
wb.save('Imprimir.xlsx')

# Clean up the temporary files from your computer
for index in range(len(rows_data)):
    row = 4 + index
    if os.path.exists(f"temp_barcode_{row}.png"):
        os.remove(f"temp_barcode_{row}.png")

print("Feito! Planilha 'Imprimir' gerada.")